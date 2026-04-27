from openpyxl import load_workbook
from openpyxl.styles import Border, Side
wb = load_workbook("gii.xlsx")
ws = wb.active

thin_border = Border(
    left = Side(style="thin", color="000000"),
    right = Side(style="thin", color="000000"),
    top = Side(style="thin", color="000000"),
    bottom = Side(style="thin", color="000000")
)

insert_position = 0
for row in ws.iter_rows():
    if row[1].value == "商品名称":
        insert_position = row[1].row+1
        break

ranges_to_move = []
# 列举插入行后面的合并单元格
for merged in ws.merged_cells.ranges:
    if merged.min_row > insert_position:
        ranges_to_move.append(merged)
# 取消合并
for merged in ranges_to_move:
    ws.unmerge_cells(str(merged))
# 插入行
ws.insert_rows(insert_position, 3)
# 将合并单元格位置下移并合并
for merged in ranges_to_move:
    new_min_row = merged.min_row + 3
    new_max_row = merged.max_row + 3

    ws.merge_cells(
        start_row = new_min_row,
        start_column = merged.min_col,
        end_row = new_max_row,
        end_column = merged.max_col
    )
# 给插入行添加边框
for row in ws.iter_rows(min_row=insert_position, max_row=insert_position+3):
    for cell in row:
        cell.border = thin_border
wb.save("new_gii.xlsx")