from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook("gii.xlsx")
ws = wb.active
oooooo = list(ws.merged_cells.ranges)

for cr in old_ranges:
    min_row, min_col, max_row, max_col = cr.min_row, cr.min_col, cr.max_row, cr.max_col
    print(f"{min_row},{min_col},{max_row}, {max_col}")