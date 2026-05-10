from datetime import date
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

class ExcelManager():
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = self.load()
        self.ws1 = self.wb.active
        self.ws1.title = str(date.today())
        self.ws2 = self.wb.create_sheet("赠品" + str(date.today()))

    def load(self):
        return load_workbook(self.file_path)
    
    def write_header(self, header_dict: dict):
        """填写表格头部信息"""
        for key in header_dict.keys():
             for rows in self.ws1.iter_rows():
                  for cell in rows:
                       if key == cell.value:
                            cell.value = header_dict[key]

    def write_body(self, product_dict: dict):
        """填写body —— 产品信息"""
        # 取得插入行数（产品数量）、第一行位置、最后一行位置
        dict_len = len(product_dict)
        first_row, first_column = self.locate_body_position()
        last_row = first_row + dict_len - 1

        # 记录body下方的合并单元格旧位置，按body行数计j出新位置
        merged_info = self.find_merged_cells(first_row, dict_len-1)
        # 拆分单元格
        self.unmerge_cells(merged_info)
        # 插入空白行
        if dict_len > 1:
            self.ws1.insert_rows(idx=first_row+1, amount=dict_len-1)
        # 产品信息写入表格
        self.write_product_messages(first_row, dict_len-1, product_dict)
        # 合并单元格
        self.merge_cells(merged_info)
        # 设置单元格边框
        self.set_border_alignment(first_row, first_row + dict_len, 1, self.ws1.max_column)

        return
    
    def locate_body_position(self):
        """找到body的插入位置"""
        target_cell = None
        # 找到body的首行
        for row in self.ws1.iter_rows():
            for cell in row:
                if cell.value == "prod_row1":
                    target_cell = cell
        if target_cell:
            return target_cell.row, target_cell.column

    def find_merged_cells(self, body_first_row:int, append_row_count:int) -> list:
        """将所有在body下方的合并单元格的位置记录在列表里面，后面添加产品后更新位置重新合并"""
        merged_info = []
        merged_ranges = self.ws1.merged_cells

        for i in merged_ranges:
            if i.max_row < body_first_row:
                continue
            info = {
                "old_coord": i.coord,
                "min_row": i.min_row + append_row_count,
                "max_row": i.max_row + append_row_count,
                "min_col": i.min_col,
                "max_col": i.max_col,
            }
            merged_info.append(info)
        return merged_info

    def unmerge_cells(self, merge_list: list):
        """拆分单元格"""
        for info in merge_list:
            self.ws1.unmerge_cells(info["old_coord"])

    def write_product_messages(self, body_first_row:int, append_row_count:int, product_dict:dict):
        """产品信息写入表格"""

        # 找到标题行
        title_row = body_first_row - 1
        # 将所有标题名存入字典{标题名：对应列号}
        title_names_positions = {}
        for row in self.ws1.iter_rows(min_row = title_row, max_row = title_row):
            for cell in row:
                title_names_positions[cell.value] = cell.column
        # 需要遍历存入的内容行
        rows = self.ws1.iter_rows(body_first_row, body_first_row + append_row_count)
        # 逐行逐列+产品字典类里的逐个属性写入
        for row, (_, [prod, serial]) in zip(rows, product_dict.items()):
            for cell in row:
                # 判断标题行的每个单元格，在下方写入对应数据
                title = self.ws1.cell(title_row, cell.column).value
                cell.value = self.match_prod_info(title, prod, serial)
                
    def match_prod_info(self, title, prod, serial):
        """根据标题，来返回对应产品"""
        match title:
            case _ if "序号" in title:
                result = serial
            case _ if "商品名称" in title:
                result = prod.name if prod.name else "请设置商品名称"
            case _ if "规格" in title:
                result = prod.model_No if prod.model_No else "请设置规格"
            case _ if "生产企业" in title:
                result = prod.company if prod.company else "请设置生产企业"
            case _ if "注册证" in title:
                result = prod.reg_No if prod.reg_No else "请设置注册证"
            case _ if "批号/序列号" in title:
                result = prod.serial_number if prod.serial_number else prod.batch_number
            case _ if "生产" in title:
                if prod.prod_date:
                    date = str(prod.prod_date)
                    result = f"20{date[:2]}-{date[2:4]}-{date[4:]}"
            case _ if "有效期" in title:
                if prod.expiry:
                    date = str(prod.expiry)
                    result = f"20{date[:2]}-{date[2:4]}-{date[4:]}"
            case _ if "数量" in title:
                result = prod.quantity
            case _ if "单位" in title:
                result = prod.unit if prod.unit else "请设置单位"
            case _ if "储存条件" in title:
                result = "常温"
            case _ if "单价" in title:
                result = prod.unit_price if prod.unit_price else "请设置单价"
            case _ if "金额" in title:
                result = prod.amount
            case _:
                result = ""
        return result

    def merge_cells(self, merge_list: list):
        """合并单元格"""
        for info in merge_list:
            self.ws1.merge_cells(start_row=info["min_row"], start_column=info["min_col"],
                                 end_row=info["max_row"], end_column=info["max_col"])

    def set_border_alignment(self, start_row: int, end_row: int, start_col: int, end_col: int):
        """设置单元格边框、内容水平垂直居中"""
        thin_border = Border(
            left    = Side(style='thin', color='000000'),
            right   = Side(style='thin', color='000000'),
            top     = Side(style='thin', color='000000'),
            bottom  = Side(style='thin', color='000000')
        )
        for i in range(start_row, end_row):
            for j in range(start_col,end_col+1):
                cell = self.ws1.cell(row=i, column=j)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def save(self, file_name):
        self.wb.save(Path.home() / f"Desktop/{file_name}.xlsx")

